# coding:utf-8
import time
import xlrd, os
from common.logger import Log
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font, colors
from common import read_config


class ReadExcel:
    """读取excel数据，返回[{},{}...]"""

    def __init__(self, excel_path, sheet_name="Sheet1"):
        self.data = xlrd.open_workbook(excel_path)
        self.table = self.data.sheet_by_name(sheet_name)
        self.keys = self.table.row_values(0)  # 获取第一行作为key值
        self.rowNum = self.table.nrows  # 获取总行数
        self.colNum = self.table.ncols  # 获取总列数
        self.log = Log()

    def dict_data(self, make=True):
        if self.rowNum <= 1:
            self.log.error("总行数小于1")
        else:
            r = []
            j = 1
            if make:
                for i in list(range(self.rowNum - 1)):  # 去掉行首 self.rowNum - 1
                    s = {'rowNum': i + 2}
                    values = self.table.row_values(j)  # 从第二行取对应values值
                    for x in list(range(self.colNum)):
                        s[self.keys[x]] = values[x]
                    r.append(s)
                    j += 1
                return r  # 返回list包含的dict数据
            else:
                for i in list(range(self.rowNum - 1)):  # 去掉行首 self.rowNum - 1
                    s = {}
                    values = self.table.row_values(j)  # 从第二行取对应values值
                    s[self.keys[1]] = values[1]
                    r.append(s)
                    j += 1
                return r  # 返回list包含的dict数据


def copy_excel(case_path, report_path, sheet_num=0):
    """复制测试用例到report_path"""
    wb2 = openpyxl.Workbook()
    wb2.save(report_path)  # 在设置的路径下创建一个excel文件
    # 读取数据
    wb1 = openpyxl.load_workbook(case_path)
    wb2 = openpyxl.load_workbook(report_path)
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[sheet_num]]  # 获取第一个sheet页
    sheet2 = wb2[sheets2[0]]
    max_row = sheet1.max_row  # 最大行数
    max_column = sheet1.max_column  # 最大列数

    for m in list(range(1, max_row + 1)):
        for n in list(range(97, 97 + max_column)):  # chr(97)='a'
            n = chr(n)  # ASCII字符,excel文件的列 a b c
            i = '%s%d' % (n, m)  # 单元格编号
            cell1 = sheet1[i].value  # 获取测试用例单元格数据
            sheet2[i].value = cell1  # 赋值到测试结果单元格

    wb2.save(report_path)  # 保存数据
    wb1.close()  # 关闭excel
    wb2.close()


class Write_excel(object):
    """写入excel数据"""

    def __init__(self, filename, title='Sheet1'):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        sheets = self.wb.sheetnames
        if title not in sheets:  # 如果没有指定的sheet页，就修改当前sheet名称
            self.ws = self.wb.active
            self.ws.title = title  # 修改sheet名称
            self.wb.save(filename)
        self.ws = self.wb[title]

    def write(self, row_n, col_n, value):
        """写入数据，如(2,3，"hello"),第二行第三列写入数据"hello\""""
        # 判断值为错误时添加字体样式
        if value in ['fail', 'error'] or col_n == 13:
            ft = Font(color=colors.RED, size=12, bold=True)
            self.ws.cell(row_n, col_n).font = ft
        if value == 'pass':
            ft = Font(color=colors.GREEN, size=12, bold=True)
            self.ws.cell(row_n, col_n).font = ft
        else:
            ft = Font(color=colors.BLACK, size=14, bold=True, family=True)
            self.ws.cell(1, col_n).font = ft
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)


def write_result(result, filename="result.xlsx"):
    """写入指定数据"""
    row_nub = result['rowNum']  # 返回结果的行数row_nub
    # 结果写入excel中
    wt = Write_excel(filename)
    wt.write(row_nub, 11, result['status_code'])  # 写入返回状态码status_code,第8列
    wt.write(row_nub, 13, result['times'])  # 耗时
    wt.write(row_nub, 14, result['error'])  # 状态码非200时或系统状态码异常时的返回信息
    wt.write(row_nub, 16, result['result'])  # 结果
    wt.write(row_nub, 17, result['msg'])  # 抛异常


if __name__ == "__main__":
    copy_excel('D:\Interface_framework_Beauty\data\demo_api.xlsx', 'D:\Interface_framework_Beauty\\report\\result.xlsx')
